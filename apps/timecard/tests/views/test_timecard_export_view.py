import urllib
from http import HTTPStatus
from io import BytesIO

import openpyxl
from django.urls import reverse

from ..base import BaseTestCase


class TestTimeCardExportView(BaseTestCase):
    url = reverse("timecard:timecard_export")

    def test_get_not_login(self):
        """
        ログイン前に画面にアクセスする
        ログイン画面にリダイレクトされることを確認
        :return:
        """
        super().base_test_get_not_login()

    def test_export_file_not_param(self):
        """
        クエリパラメーターなしでアクセスする
        打刻一覧画面にリダイレクトされることを確認
        :return:
        """
        response = self.client.get(self.url, follow=True)
        self.assertIn(reverse("timecard:timecard_monthly_report"), response.redirect_chain[0][0])
        self.assertEqual(HTTPStatus.FOUND.value, response.redirect_chain[0][1])

    def test_export_file_error_param(self):
        """
        不正なクエリパラメーターでアクセスする
        打刻一覧画面にリダイレクトされることを確認
        :return:
        """
        response = self.client.get(self.url, {"a": "a"}, follow=True)
        self.assertIn(reverse("timecard:timecard_monthly_report"), response.redirect_chain[0][0])
        self.assertEqual(HTTPStatus.FOUND.value, response.redirect_chain[0][1])

    def test_filename_normally_param(self):
        """
        正常なクエリパラメーターでアクセスする（key: month, value: YYYYMM）
        ダウンロードしたExcelのファイル名を確認
        :return:
        """
        response = self.client.get(self.url, {"month": "202305"})
        self.assertTrue(response.has_header("Content-Disposition"))

        response_header = response.headers.get("Content-Disposition")
        self.assertIn("勤怠_テスト2_2023年05月.xlsx", urllib.parse.unquote(response_header))

    def test_content_no_record(self):
        """
        データが0件
        ダウンロードしたExcelの中身を確認する
        :return:
        """
        ws = self._get_ws_by_response(month="202305")

        self.assertEqual("勤務報告書　2023年05月", ws["A2"].value)
        self.assertEqual("氏名　テスト2", ws["A4"].value)

        self._assert_header(ws[5])

        for row_num in range(6, 36):
            row = ws[row_num]
            self.assertIsNone(row[3].value)
            self.assertIsNone(row[4].value)
            self.assertIsNone(row[5].value)
            self.assertIsNone(row[6].value)

        self.assertEqual("憲法記念日", ws["H8"].value)
        self.assertEqual("みどりの日", ws["H9"].value)
        self.assertEqual("こどもの日", ws["H10"].value)

    def test_content_record(self):
        """
        データが1件
        ダウンロードしたExcelの中身を確認する
        :return:
        """
        ws = self._get_ws_by_response(month="202301")

        self.assertEqual("勤務報告書　2023年01月", ws["A2"].value)

        self.assertEqual("09:00", ws["D7"].value)
        self.assertEqual("19:00", ws["E7"].value)
        self.assertEqual("12:00", ws["F7"].value)
        self.assertEqual("13:00", ws["G7"].value)

    def _get_ws_by_response(self, **get_params):
        """
        ダウンロードしたExcelのシート名を確認
        :return::class:`openpyxl.worksheet.worksheet.Worksheet`
        """
        response = self.client.get(self.url, get_params)
        self.assertTrue(response.has_header("Content-Disposition"))

        wb = openpyxl.load_workbook(BytesIO(response.content))
        self.assertEqual(1, len(wb.sheetnames))
        self.assertEqual("一覧", wb.sheetnames[0])

        return wb["一覧"]

    def _assert_header(self, header_row):
        """
        ヘッダーを確認する
        :param header_row: ヘッダーの行
        :return:
        """
        self.assertEqual("日付", header_row[0].value)
        self.assertEqual("曜日", header_row[1].value)
        self.assertEqual("MergedCell", type(header_row[2]).__name__)
        self.assertEqual("出勤", header_row[3].value)
        self.assertEqual("退勤", header_row[4].value)
        self.assertEqual("休憩開始", header_row[5].value)
        self.assertEqual("休憩終了", header_row[6].value)
        self.assertEqual("備考", header_row[7].value)
