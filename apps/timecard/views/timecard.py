import logging
import re
import urllib
from datetime import datetime, time

import jpholiday
import openpyxl
from dateutil.relativedelta import relativedelta
from django.db import transaction
from django.db.models import Q
from django.db.models.functions import TruncMonth
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.template.response import TemplateResponse
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views.generic import ListView
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from apps.accounts.models import User
from apps.timecard.forms import (TimeCardFormSet, TimeCardSearchForm,
                                 UploadFileForm)
from apps.timecard.models import TimeCard, TimeCardSummary

from .base import (ExcelHandleView, SuperuserPermissionView, TemplateView,
                   TimeCardBaseMonthlyReportView, get_DOW,
                   get_toast_msg_by_session, timedelta2str)


class TimeCardMonthlyReportView(TimeCardBaseMonthlyReportView):
    template_name = "material-dashboard-master/pages/new_list.html"
    session_month_format = "%Y-%m-%d %H:%M:%S%z"

    def _get_EOM_by_url(self):
        if not self.request.GET.get("month") and self.request.session.get("month"):
            return datetime.strptime(self.request.session["month"], self.session_month_format)
        return super()._get_EOM_by_url()

    def get_context_data(self):
        context = super().get_context_data()
        context["search_form"] = self._get_search_form()
        context["promote_err_msg"] = self._get_promote_err_msg_by_session(self.request.session)
        context["can_edit"] = self._get_can_edit()

        return context

    def _get_search_form(self):
        kwargs = {}
        kwargs["initial"] = {"month": self.EOM_by_url.strftime("%Y-%m")}
        self.request.session["month"] = self.EOM_by_url.strftime(self.session_month_format)
        return TimeCardSearchForm(**kwargs)

    def _get_promote_err_msg_by_session(self, session):
        YYYYMM = self.EOM_by_url.strftime("%Y%m")
        if session.get("promote_err_month_dict") and session["promote_err_month_dict"].get(YYYYMM):
            promote_err_dict = session["promote_err_month_dict"][YYYYMM]
            return ["{}日：{}".format(day, err_msg) for day, err_msg in promote_err_dict.items()]

    def _get_can_edit(self):
        today = timezone.datetime.today().astimezone(timezone.get_default_timezone())
        first_day_next_month = today + relativedelta(months=1, day=1)
        return self.EOM_by_url < first_day_next_month

    def _promote_process(self):
        url = reverse("timecard:timecard_monthly_report") + "?month=" + self.EOM_by_url.strftime("%Y%m")
        monthly_stamps_qs = self.get_queryset()

        if not monthly_stamps_qs.exists():
            self.request.session["warning"] = "未打刻のため申請できません"
            return redirect(url)
        elif monthly_stamps_qs.exclude(state=TimeCard.State.NEW):
            self.request.session["error"] = "すでに申請済みです"
            return redirect(url)

        promote_err_dict = {}
        if self._is_valid_stamps_qs(monthly_stamps_qs, promote_err_dict):
            monthly_stamps_qs.update(state=TimeCard.State.PROCESSING)
            self.request.session["success"] = "ステータスを{}に更新しました".format(TimeCard.State.PROCESSING.label)
            return redirect(url)

        self.request.session["warning"] = "入力時刻にエラーがあるため申請できません"
        self.request.session["promote_err_month_dict"] = {self.EOM_by_url.strftime("%Y%m"): promote_err_dict}
        return redirect(url)

    def _is_valid_stamps_qs(self, monthly_stamps_qs, promote_err_dict):
        work_days_list = self._get_work_days_by_qs(monthly_stamps_qs)

        for work_day in work_days_list:
            start_work, end_work, enter_break, end_break = self._get_daily_stamps_info(monthly_stamps_qs, work_day)

            if not (start_work and end_work):
                promote_err_dict[work_day] = TimeCardFormSet.ERR_MSG_NEED_WORK_TIME
                continue

            elif not (start_work < end_work):
                promote_err_dict[work_day] = TimeCardFormSet.ERR_MSG_WORK_TIME
                continue

            if enter_break == end_break == "":
                continue

            if not (enter_break and end_break):
                promote_err_dict[work_day] = TimeCardFormSet.ERR_MSG_NEED_BREAK_TIME
                continue

            elif not (enter_break < end_break):
                promote_err_dict[work_day] = TimeCardFormSet.ERR_MSG_BREAK_TIME
                continue

            if not (start_work <= enter_break) or not (end_break <= end_work):
                promote_err_dict[work_day] = TimeCardFormSet.ERR_MSG_BREAK_TIME_OUT_OF_RANGE
                continue

        return promote_err_dict == {}


class TimeCardExportView(TimeCardBaseMonthlyReportView, ExcelHandleView):
    def get(self, request, *args, **kwargs):
        self.EOM_by_url = super()._get_EOM_by_url(False)

        if self.EOM_by_url is None:
            return redirect(reverse("timecard:timecard_monthly_report"))

        wb = self._create_wb()
        filename = "タイムカード_" + self.EOM_by_url.strftime("%Y{0}%m{1}").format(*"年月") + ".xlsx"
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename={}".format(urllib.parse.quote(filename))
        wb.save(response)

        return response

    def _create_wb(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.title = self.SHEET_TITLE

        self._write_ws(ws)
        self._edit_appearance_ws(ws)
        self._adjust_col_width_ws(ws)

        return wb

    def _write_ws(self, ws):
        self._write_header(ws)

        day_count = self.EOM_by_url.day
        monthly_stamps_qs = self.get_queryset()
        for day_index in range(day_count):
            day = day_index + 1
            target_date = self.EOM_by_url + relativedelta(day=day)
            holiday_name = jpholiday.is_holiday_name(target_date)

            DOW = self._get_DOW(day)
            day_kind = self._get_day_kind(DOW, day)
            start_time, end_time, enter_break, end_break = self._get_daily_stamps_info(monthly_stamps_qs, day)

            ws.append(
                [
                    day,
                    DOW,
                    day_kind,
                    self._local_time(start_time),
                    self._local_time(end_time),
                    self._local_time(enter_break),
                    self._local_time(end_break),
                    holiday_name,
                ]
            )

    def _format(self, timedelta):
        if type(timedelta) == str:
            return timedelta

        pattern = r"^((0?|1)[0-9]|2[0-3]):[0-5][0-9]$"
        return re.compile(pattern).match(str(timedelta)).group()

    def _local_time(self, stamped_time):
        if stamped_time:
            return timezone.localtime(stamped_time).strftime("%H:%M")

        return ""

    def _write_header(self, ws):
        ws[self.TITLE_CELL] = self.SHEET_HEADLINE.format(self.EOM_by_url.strftime("%Y{0}%m{1}").format(*"年月"))
        ws[self.USER_NAME_CELL] = self.SHEET_USER_NAME.format(self.request.user.name)

        ws.append(self.SHEET_HEADER)


class TimeCardEditView(TemplateView):
    template_name = "material-dashboard-master/pages/form.html"

    def dispatch(self, request, *args, **kwargs):
        self.date_by_url = self._get_date_by_url()

        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        if not self._can_edit():
            return render(request, self.template_name)

        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        formset = self._get_formset(request.POST)
        if self._has_promoted_stamps(formset):
            request.session["error"] = "申請中のため編集できません"
            return render(request, self.template_name)

        if not formset.is_valid():
            if not self.date_by_url:
                return redirect(reverse("timecard:timecard_monthly_report"))

            context = self.get_context_data(formset)
            return render(request, self.template_name, context)

        delete_data = len([key for key in formset.data.keys() if "DELETE" in key]) > 0
        saved_form_list = formset.save()

        if saved_form_list is False:
            request.session["error"] = "更新に失敗しました"
            return render(request, self.template_name)

        elif delete_data or len(saved_form_list) > 0:
            request.session["success"] = "更新しました"
            self._delete_promote_err_msg_by_session(request.session)
            return render(request, self.template_name)

        return redirect(reverse("timecard:timecard_monthly_report"))

    def _get_date_by_url(self):
        param = self.request.GET.get("date")
        if not param:
            return

        try:
            date_by_url = timezone.datetime.strptime(param, "%Y%m%d").astimezone(timezone.get_default_timezone())
        except:
            return

        return date_by_url

    def get_context_data(self, formset=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context["formset"] = formset or self._get_formset(date=self.date_by_url, user=self.request.user)
        context["edit_date"] = self.date_by_url.date()
        context["DOW"] = get_DOW(self.date_by_url.date())
        return context

    def _get_formset(self, *args, **kwargs):
        formset = TimeCardFormSet(*args, **kwargs)
        for form in formset:
            form.instance.user = self.request.user

        return formset

    def _delete_promote_err_msg_by_session(self, session):
        if not self.date_by_url:
            return

        if session.get("promote_err_month_dict") and session["promote_err_month_dict"].get(
            self.date_by_url.strftime("%Y%m")
        ):
            session["promote_err_month_dict"][self.date_by_url.strftime("%Y%m")].pop(str(self.date_by_url.day), None)

    def _can_edit(self):
        next_month = timezone.datetime.today().astimezone(timezone.get_default_timezone()).replace(
            hour=0, minute=0, second=0, microsecond=0
        ) + relativedelta(day=1, months=1)
        if not self.date_by_url:
            self.request.session["error"] = "不正な操作を検知しました"
            return

        elif self.date_by_url >= next_month:
            self.request.session["error"] = "来月以降の情報は編集できません"
            return

        daily_stamps_qs = TimeCard.objects.filter(
            stamped_time__gte=(self.date_by_url + relativedelta(day=1)),
            stamped_time__lt=(self.date_by_url + relativedelta(months=1, day=1)),
            user=self.request.user,
        ).exclude(state=TimeCard.State.NEW)
        if daily_stamps_qs.exists():
            self.request.session["error"] = "{}および{}の情報は更新できません".format(
                TimeCard.State.PROCESSING.label, TimeCard.State.APPROVED.label
            )
            return

        return True

    def _has_promoted_stamps(self, formset):
        for form in formset:
            if form.instance.id is None:
                return

            obj = TimeCard.objects.get(id=form.instance.id)
            return obj and obj.state != TimeCard.State.NEW


class TimeCardImportView(ExcelHandleView):
    template_name = "material-dashboard-master/pages/upload.html"
    logger = logging.getLogger(__name__)

    ERR_MSG_HEADER = "エラー内容"
    ERR_FORMAT = "フォーマットはHH:MMで入力してください。"

    def __init__(self):
        super().__init__()
        self.err_cell_list = []
        self.row_err_msg_dict = {}
        self.EOM_by_ws = None

    def get(self, *args, **kwargs):
        return TemplateResponse(self.request, self.template_name, {"upload_form": UploadFileForm()})

    def post(self, request, *args, **kwargs):
        upload_form = self._get_upload_form()
        if not (upload_form.is_valid()):
            return render(request, self.template_name, {"upload_form": upload_form})

        wb = openpyxl.load_workbook(request.FILES["file"])

        if self._can_import_wb(wb):
            self._import_data_by_wb(wb)
            return render(request, self.template_name)

        if request.session.get("error"):
            return render(request, self.template_name)

        err_response = self._create_err_response(wb)
        return err_response

    def _get_upload_form(self):
        kwargs = {"data": self.request.POST, "files": self.request.FILES}
        return UploadFileForm(**kwargs)

    def _can_import_wb(self, wb):
        try:
            ws = wb[self.SHEET_TITLE]

            if self._invalid_ws_layout(ws):
                return

            if self._exist_promoted_stamps():
                return

            if self._invalid_ws_data(ws):
                return

            return True

        except Exception as e:
            self.request.session["error"] = "取込失敗しました"
            self.logger.error(f"{e}", exc_info=True)
            return

    def _invalid_ws_layout(self, ws):
        user_name = ws[self.USER_NAME_CELL].value[3:]
        if user_name != self.request.user.name:
            self.request.session["error"] = "不正なシートのため取込できません"
            return True

        self.EOM_by_ws = self._get_EOM_by_ws(ws)

        # シートの月末日が正しいかチェックする
        if self.EOM_by_ws.day != ws.cell(row=ws.max_row, column=1).value:
            self.request.session["error"] = "不正なシートのため取込できません"
            return True

        next_month = timezone.datetime.today().astimezone(timezone.get_default_timezone()).replace(
            hour=0, minute=0, second=0, microsecond=0
        ) + relativedelta(day=1, months=1)

        if self.EOM_by_ws >= next_month:
            self.request.session["error"] = "来月以降の情報は取込できません"
            return True

        return False

    def _invalid_ws_data(self, ws):
        empty_row_count = 0
        valid_row_count = 0
        # 1日の行から末日の行までループする
        for row_num in range(self.HEADER_ROW_NUM + 1, ws.max_row + 1):
            row = ws[row_num]

            start_work_cell = row[self.START_WORK_COL_NUM - 1]
            end_work_cell = row[self.END_WORK_COL_NUM - 1]
            enter_break_cell = row[self.ENTER_BREAK_COL_NUM - 1]
            end_break_cell = row[self.END_BREAK_COL_NUM - 1]

            if start_work_cell.value is end_work_cell.value is enter_break_cell.value is end_break_cell.value is None:
                empty_row_count += 1
                continue

            if self._exist_format_err(start_work_cell, end_work_cell, enter_break_cell, end_break_cell):
                self.row_err_msg_dict[row_num] = self.ERR_FORMAT
                continue

            if self._invalid_work_time(start_work_cell, end_work_cell):
                continue

            if enter_break_cell.value is end_break_cell.value is None:
                valid_row_count += 1
                continue

            if self._invalid_break_time(start_work_cell, end_work_cell, enter_break_cell, end_break_cell):
                continue

            valid_row_count += 1

        day_count = self.EOM_by_ws.day
        if day_count == empty_row_count:
            # 未入力の場合、取込した月の情報が全て削除されるためエラーにする
            self.request.session["error"] = "未入力のため取込できません"
            return

        return day_count != (empty_row_count + valid_row_count)

    def _reset_color_cells(self, ws):
        for col in ws.columns:
            if col[0].column <= self.DAY_KIND_COL_NUM:
                # '日付'、'曜日'、'区分'の列の色は変更しない
                continue

            for cell in col:
                cell.fill = PatternFill(fgColor=self.WHITE)

    def _reset_err_msg_col(self, ws):
        ws.delete_cols(self.ERR_MSG_COL_NUM)
        ws.cell(row=self.HEADER_ROW_NUM, column=self.ERR_MSG_COL_NUM).value = self.ERR_MSG_HEADER
        ws.cell(row=self.HEADER_ROW_NUM, column=self.ERR_MSG_COL_NUM).font = Font(name=self.FONT_NAME, size=9)

    def _write_err_msg_ws(self, err_msg, err_msg_cell, *err_cells):
        err_msg_cell.value = err_msg
        err_msg_cell.font = Font(name=self.FONT_NAME, size=9)
        for err_cell in err_cells:
            err_cell.fill = PatternFill(patternType="solid", fgColor=self.YELLOW)

    def _write_borders_ws(self, ws):
        side = Side(style="thin", color=self.BLACK)
        for row in ws:
            row_num = row[0].row

            if row_num < self.HEADER_ROW_NUM:
                continue

            # ヘッダーから末日の行までループする
            for cell in row:
                ws[cell.coordinate].border = Border(top=side, bottom=side, left=side, right=side)
                ws[cell.coordinate].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

    def _cell_value2time(self, cell_value):
        if cell_value is None or type(cell_value) == time:
            return cell_value

        try:
            return datetime.strptime(cell_value, "%H:%M").time()
        except:
            return False

    def _get_EOM_by_ws(self, ws):
        value = ws[self.TITLE_CELL].value[6:]
        month_str = value.replace("年", "").replace("月", "") + "01"
        import_month = datetime.strptime(month_str, "%Y%m%d").astimezone(timezone.get_default_timezone())
        return import_month + relativedelta(months=1, day=1) - relativedelta(days=1)

    @transaction.atomic
    def _import_data_by_wb(self, wb):
        try:
            with transaction.atomic():
                ws = wb[self.SHEET_TITLE]

                self._get_queryset().delete()

                # 1日の行から末日の行までループする
                stamp_list = []
                for row_num in range(self.HEADER_ROW_NUM + 1, ws.max_row + 1):
                    row = ws[row_num]
                    day = row[0].value

                    stamped_date = self.EOM_by_ws + relativedelta(day=day)
                    col_num_kind = {
                        self.START_WORK_COL_NUM: TimeCard.Kind.IN,
                        self.END_WORK_COL_NUM: TimeCard.Kind.OUT,
                        self.ENTER_BREAK_COL_NUM: TimeCard.Kind.ENTER_BREAK,
                        self.END_BREAK_COL_NUM: TimeCard.Kind.END_BREAK,
                    }

                    for col_num in col_num_kind:
                        cell_value = row[col_num - 1].value

                        if cell_value:
                            stamp_kind = col_num_kind[col_num]
                            stamped_time = self._make_stamped_time(stamped_date, cell_value)
                            stamp_list.append([stamp_kind, stamped_time])

                if stamp_list:
                    for stamp in stamp_list:
                        TimeCard.objects.create(kind=stamp[0], stamped_time=stamp[1], user=self.request.user)

                self.request.session["success"] = "{}の取込が成功しました".format(
                    self.EOM_by_ws.strftime("%Y{0}%m{1}").format(*"年月")
                )

                if self.request.session.get("promote_err_month_dict"):
                    self.request.session["promote_err_month_dict"].pop(self.EOM_by_ws.strftime("%Y%m"))

        except Exception as e:
            self.logger.error(f"{e}", exc_info=True)

    def _make_stamped_time(self, stamped_day, cell_value):
        if type(cell_value) != time:
            cell_value = datetime.strptime(cell_value, "%H:%M").time()
        return stamped_day + relativedelta(hours=cell_value.hour, minutes=cell_value.minute)

    def _create_err_response(self, wb):
        ws = wb[self.SHEET_TITLE]
        self._reset_color_cells(ws)
        self._reset_err_msg_col(ws)

        self._set_err_color()
        self._set_err_msg(ws)

        self._edit_appearance_ws(ws)
        self._adjust_col_width_ws(ws)

        filename = "エラーレポート_" + timezone.datetime.now().strftime("%Y%m%d%H%M") + ".xlsx"
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename={}".format(urllib.parse.quote(filename))
        wb.save(response)

        return response

    def _exist_promoted_stamps(self):
        monthly_stamps_qs = self._get_queryset()
        if monthly_stamps_qs.exclude(state=TimeCard.State.NEW).exists():
            self.request.session["error"] = "{}および{}の情報は更新できません".format(
                TimeCard.State.PROCESSING.label, TimeCard.State.APPROVED.label
            )

        return monthly_stamps_qs.exclude(state=TimeCard.State.NEW).exists()

    def _get_queryset(self):
        monthly_stamps_qs = TimeCard.objects.filter(
            stamped_time__gte=(self.EOM_by_ws + relativedelta(day=1)),
            stamped_time__lt=(self.EOM_by_ws + relativedelta(months=1, day=1)),
            user=self.request.user,
        )

        return monthly_stamps_qs

    def _exist_format_err(self, *cells):
        format_err = False
        for cell in cells:
            if self._cell_value2time(cell.value) is False:
                self.err_cell_list.append(cell)
                format_err = True

        return format_err

    def _set_err_color(self):
        for err_cell in self.err_cell_list:
            err_cell.fill = PatternFill(patternType="solid", fgColor=self.YELLOW)

    def _set_err_msg(self, ws):
        for row_num, err_msg in self.row_err_msg_dict.items():
            ws[row_num][self.ERR_MSG_COL_NUM - 1].value = err_msg
            ws[row_num][self.ERR_MSG_COL_NUM - 1].font = Font(name=self.FONT_NAME, size=9)

    def _invalid_work_time(self, start_work_cell, end_work_cell):
        start_work = self._cell_value2time(start_work_cell.value)
        end_work = self._cell_value2time(end_work_cell.value)
        row_num = start_work_cell.row

        if (start_work is None) ^ (end_work is None):
            self.err_cell_list.extend([start_work_cell, end_work_cell])
            self.row_err_msg_dict[row_num] = TimeCardFormSet.ERR_MSG_NEED_WORK_TIME
            return True

        elif end_work < start_work:
            self.err_cell_list.extend([start_work_cell, end_work_cell])
            self.row_err_msg_dict[row_num] = TimeCardFormSet.ERR_MSG_WORK_TIME
            return True

        return False

    def _invalid_break_time(self, start_work_cell, end_work_cell, enter_break_cell, end_break_cell):
        start_work = self._cell_value2time(start_work_cell.value)
        end_work = self._cell_value2time(end_work_cell.value)
        enter_break = self._cell_value2time(enter_break_cell.value)
        end_break = self._cell_value2time(end_break_cell.value)
        row_num = start_work_cell.row

        if (enter_break is None) ^ (end_break is None):
            self.err_cell_list.extend([enter_break_cell, end_break_cell])
            self.row_err_msg_dict[row_num] = TimeCardFormSet.ERR_MSG_NEED_BREAK_TIME
            return True

        elif end_break < enter_break:
            self.err_cell_list.extend([enter_break_cell, end_break_cell])
            self.row_err_msg_dict[row_num] = TimeCardFormSet.ERR_MSG_BREAK_TIME
            return True

        elif enter_break < start_work or end_work < end_break:
            self.err_cell_list.extend([enter_break_cell, end_break_cell])
            self.row_err_msg_dict[row_num] = TimeCardFormSet.ERR_MSG_BREAK_TIME_OUT_OF_RANGE
            return True

        return False


class TimeCardProcessMonthListView(SuperuserPermissionView, ListView):
    context_object_name = "month_list"
    template_name = "material-dashboard-master/pages/tables_process.html"

    def get_queryset(self):
        state_process_month_qs = (
            TimeCard.objects.filter(~Q(user=self.request.user), state=TimeCard.State.PROCESSING)
            .values("user", "stamped_time")
            .annotate(month=TruncMonth("stamped_time"))
            .values("user", "month")
            .order_by("user", "-month")
            .distinct()
        )

        return state_process_month_qs

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data()
        context.update(**get_toast_msg_by_session(self.request.session))

        for month in context["month_list"]:
            user = User.objects.get(pk=month["user"])
            month["user_name"] = user.name
            month["date_str"] = month["month"].strftime("%Y{0}%m{1}").format(*"年月")
            month["param"] = "?month={}&user={}".format(month["month"].strftime("%Y%m"), user.id)

        return context



class TimeCardApprovedMonthListView(SuperuserPermissionView, ListView):
    template_name = "material-dashboard-master/pages/tables_approved.html"

    def get(self, request, *args, **kwargs):
        display_month = self._get_month_str_by_url() or self.request.session.get('approved_month') or (timezone.datetime.today() - relativedelta(months=1)).strftime("%Y%m")
        self.request.session['approved_month'] = display_month
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        summary_qs = TimeCardSummary.objects.filter(~Q(user=self.request.user), month=self.request.session['approved_month'])

        return summary_qs

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data()
        context.update(**get_toast_msg_by_session(self.request.session))
        context["search_form"] = self._get_search_form()

        return context

    def _get_search_form(self):
        kwargs = {}
        kwargs["initial"] = {"month": self.request.session['approved_month'][:4] + '-' + self.request.session['approved_month'][4:]}
        return TimeCardSearchForm(**kwargs)

    def _get_month_str_by_url(self):
        param = self.request.GET.get("month")
        if not param:
            return
        try:
            timezone.datetime.strptime(param + "01", "%Y%m%d")
            return param
        except:
            return


class TimeCardApprovedMonthlyReportView(SuperuserPermissionView, TimeCardBaseMonthlyReportView):
    template_name = "material-dashboard-master/pages/approved_list.html"
    url = reverse_lazy("timecard:timecard_approved_month_list")
    DISPLAY_STATE = TimeCard.State.APPROVED

    def get(self, request, *args, **kwargs):
        self.user = self._get_user_by_url(request)

        self.EOM_by_url = self._get_EOM_by_url()
        if self.user is None or self.EOM_by_url is None:
            request.session["error"] = "不正な操作を検知しました"
            return redirect(self.url)

        if not self.get_queryset().exists():
            request.session["error"] = "勤怠情報が存在しません"
            return redirect(self.url)

        return super().get(request, *args, **kwargs)

    def _get_user_by_url(self, request):
        param = request.GET.get("user")
        if param:
            try:
                if param == request.user.id:
                    return

                return User.objects.get(pk=param)
            except:
                pass

        return

    def _get_EOM_by_url(self):
        return super()._get_EOM_by_url(False)

    def get_context_data(self):
        context = super().get_context_data()
        context["user_name"] = self.user.name
        context["total_work_hours"] = self._get_total_work_hours()

        return context

    def get_queryset(self):
        monthly_stamps_qs = TimeCard.objects.filter(
            stamped_time__gte=(self.EOM_by_url + relativedelta(day=1)),
            stamped_time__lt=(self.EOM_by_url + relativedelta(months=1, day=1)),
            user=self.user,
            state=self.DISPLAY_STATE
        ).order_by("stamped_time")

        return monthly_stamps_qs

    def _get_total_work_hours(self):
        summary = TimeCardSummary.objects.get(user=self.user, month=self.EOM_by_url.strftime("%Y%m"))
        return summary.total_work_hours

class TimeCardProcessMonthlyReportView(TimeCardApprovedMonthlyReportView):
    template_name = "material-dashboard-master/pages/processing_list.html"
    url = reverse_lazy("timecard:timecard_process_month_list")
    DISPLAY_STATE = TimeCard.State.PROCESSING

    def get(self, request, *args, **kwargs):
        response = super().get(request, *args, **kwargs)

        if request.session.get("error"):
            return redirect(self.url)

        if request.GET.get("mode") == "demote":
            return self._demote_process()

        return response

    def _get_total_work_hours(self):
        return timedelta2str(self.total_work_hours)

    @transaction.atomic
    def approval_process(self, monthly_stamps_qs):
        try:
            with transaction.atomic():
                monthly_stamps_qs.update(state=TimeCard.State.APPROVED)
                work_days_list = self._get_work_days_by_qs(monthly_stamps_qs)
                work_days_flag = self.create_work_days_flag(work_days_list)
                self._calculation_total_work_hours_total_break_hours(monthly_stamps_qs)

                return TimeCardSummary.objects.create(
                    user=self.user,
                    work_days_flag=work_days_flag,
                    month=self.EOM_by_url.strftime("%Y%m"),
                    total_work_hours=timedelta2str(self.total_work_hours),
                    total_break_hours=timedelta2str(self.total_break_hours),
                )
        except:
            return

    def create_work_days_flag(self, work_days_list) -> int:
        work_days_flag_int = 0
        for work_day in work_days_list:
            left_shift_count = work_day - 1
            work_days_flag_int = work_days_flag_int | 1 << left_shift_count

        return work_days_flag_int

    def _promote_process(self):
        monthly_stamps_qs = self.get_queryset()

        if not monthly_stamps_qs.exists():
            self.request.session["error"] = "申請中の勤怠情報が存在しません"
            return redirect(self.url)

        if self.approval_process(monthly_stamps_qs):
            self.request.session["success"] = "承認しました"
            return redirect(self.url)

        self.request.session["error"] = "承認処理に失敗しました"
        return redirect(self.url)

    def _calculation_total_work_hours_total_break_hours(self, monthly_stamps_qs):
        self._get_monthly_report(monthly_stamps_qs)

    def get_queryset(self):
        monthly_stamps_qs = TimeCard.objects.filter(
            stamped_time__gte=(self.EOM_by_url + relativedelta(day=1)),
            stamped_time__lt=(self.EOM_by_url + relativedelta(months=1, day=1)),
            user=self.user,

        ).order_by("stamped_time")

        return monthly_stamps_qs

    def _demote_process(self):
        try:
            self.get_queryset().update(state=TimeCard.State.NEW)
            self.request.session["success"] = "差し戻しました"
        except:
            self.request.session["error"] = "差戻処理に失敗しました"

        return redirect(self.url)

    def get_context_data(self):
        context = super().get_context_data()
        context['work_days_count'] = len(self._get_work_days_by_qs(self.get_queryset()))
        return context
