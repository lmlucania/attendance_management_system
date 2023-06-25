from datetime import timedelta
import re

from dateutil.relativedelta import relativedelta
from django.contrib.auth.mixins import LoginRequiredMixin, AccessMixin
from django.db.models.functions import TruncDay
from django.utils import timezone
from django.urls import reverse_lazy
from django.views import generic
import jpholiday
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from apps.timecard.models import TimeCard


class TemplateView(LoginRequiredMixin, generic.TemplateView):
    login_url = reverse_lazy('accounts:login')

    def get_context_data(self):
        return get_toast_msg_by_session(self.request.session)


class ListView(LoginRequiredMixin, generic.ListView):
    search_form = None


class View(LoginRequiredMixin, generic.View):
    pass


class SuperuserPermissionView(AccessMixin):
    def dispatch(self, request, *args, **kwargs):
        if not (request.user.is_authenticated and request.user.is_manager):
            return self.handle_no_permission()
        return super().dispatch(request, *args, **kwargs)


class TimeCardBaseMonthlyReportView(TemplateView):
    model = TimeCard
    template_name = None

    def get(self, request, *args, **kwargs):
        self.EOM_by_url = self.get_EOM_by_url()
        if request.GET.get('mode') == 'promote':
            return self._promote_process()
        return super().get(request, *args, **kwargs)

    def get_EOM_by_url(self, allow_none_param=True):
        today = timezone.datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
        date_by_url = None

        param = self.request.GET.get('month')
        if not (param or allow_none_param):
            return
        try:
            date_by_url = timezone.datetime.strptime(param + '01', "%Y%m%d")
        except:
            pass

        EOM_by_url = (date_by_url or today) + relativedelta(months=1, day=1) - relativedelta(days=1)
        return EOM_by_url.astimezone(timezone.get_default_timezone())

    def get_context_data(self):
        monthly_stamps_qs = self.get_queryset()

        context = super().get_context_data()
        context['monthly_report'] = self._get_monthly_report(monthly_stamps_qs)
        context['state'] = self._get_state(monthly_stamps_qs)
        context['BOM'] = (self.EOM_by_url + relativedelta(day=1)).date()
        context['EOM'] = self.EOM_by_url.date()

        return context

    def get_queryset(self):
        monthly_stamps_qs = TimeCard.objects.filter(stamped_time__gte=(self.EOM_by_url + relativedelta(day=1)),
                                                    stamped_time__lt=(self.EOM_by_url + relativedelta(months=1, day=1)),
                                                    user=self.request.user).order_by('stamped_time')
        return monthly_stamps_qs

    def _get_state(self, monthly_stamps_qs):
        if monthly_stamps_qs.filter(state=TimeCard.State.APPROVED).exists():
            return TimeCard.State.APPROVED
        elif monthly_stamps_qs.filter(state=TimeCard.State.PROCESSING).exists():
            return TimeCard.State.PROCESSING

        return TimeCard.State.NEW

    def _get_monthly_report(self, monthly_stamps_qs):
        self.total_work_hours = self.total_break_hours = timedelta()

        day_count = self.EOM_by_url.day
        work_days_list = self._get_work_days_by_qs(monthly_stamps_qs)

        monthly_report = []
        for day_index in range(day_count):
            start_work = end_work = enter_break = end_break = ''
            work_hours = break_hours = timedelta()

            day = day_index + 1
            DOW = self._get_DOW(day)
            day_kind = self._get_day_kind(DOW, day)
            date = (self.EOM_by_url + relativedelta(day=day)).strftime('%m/%d') + '({})'.format(DOW)
            holiday_name = jpholiday.is_holiday_name(self.EOM_by_url + relativedelta(day=day))

            if day in work_days_list:
                start_work, end_work, enter_break, end_break = self._get_daily_stamps_info(monthly_stamps_qs, day)
                work_hours, break_hours = self._calculation_hours_daily(start_work, end_work, enter_break, end_break)

            self.total_work_hours += work_hours
            self.total_break_hours += break_hours

            monthly_report.append({'day': day, 'date': date, 'start_work': start_work, 'end_work': end_work,
                                   'enter_break': enter_break, 'end_break': end_break, 'holiday_name': holiday_name,
                                   'work_hours': self._timedelta2str(work_hours),
                                   'break_hours': self._timedelta2str(break_hours), 'day_kind': day_kind})
        return monthly_report

    def _get_daily_stamps_info(self, monthly_stamps_qs, day):
        daily_stamps_qs = monthly_stamps_qs.filter(
            stamped_time__gte=(self.EOM_by_url + relativedelta(day=day)),
            stamped_time__lt=(self.EOM_by_url + relativedelta(day=day) + relativedelta(days=1)))

        start_work_obj = daily_stamps_qs.filter(kind=TimeCard.Kind.IN).first()
        end_work_obj = daily_stamps_qs.filter(kind=TimeCard.Kind.OUT).first()
        enter_break_obj = daily_stamps_qs.filter(kind=TimeCard.Kind.ENTER_BREAK).first()
        end_break_obj = daily_stamps_qs.filter(kind=TimeCard.Kind.END_BREAK).first()

        start_work = start_work_obj.stamped_time if start_work_obj else ''
        end_work = end_work_obj.stamped_time if end_work_obj else ''
        enter_break = enter_break_obj.stamped_time if enter_break_obj else ''
        end_break = end_break_obj.stamped_time if end_break_obj else ''

        return start_work, end_work, enter_break, end_break

    def _calculation_hours_daily(self, start_work, end_work, enter_break, end_break):
        try:
            work_hours = break_hours = timedelta()
            if start_work and end_work and start_work < end_work:
                work_hours = end_work - start_work

            if enter_break and end_break and enter_break < end_break:
                break_hours = end_break - enter_break

            if break_hours > timedelta() and work_hours > break_hours:
                work_hours -= break_hours

            return work_hours, break_hours
        except:
            return timedelta(), timedelta()

    def _get_DOW(self, day):
        week = {1: '月', 2: '火', 3: '水', 4: '木', 5: '金', 6: '土', 7: '日'}
        DOW_int = (self.EOM_by_url + relativedelta(day=day)).isoweekday()
        return week[DOW_int]

    def _get_day_kind(self, DOW, day):
        date = self.EOM_by_url + relativedelta(day=day)
        if jpholiday.is_holiday(date):
            return '祝日'
        if DOW in ['土', '日']:
            return '休日'

        return '平日'

    def _get_work_days_by_qs(self, monthly_stamps_qs):
        work_days = [stamp_qs['datetime'].day for stamp_qs in monthly_stamps_qs.annotate(
            datetime=TruncDay('stamped_time')).values('datetime').order_by('datetime').distinct()]
        return work_days

    def _timedelta2str(self, timedelta):
        if type(timedelta) == str:
            return timedelta

        pattern = r'((0?|1)[0-9]|2[0-3]):[0-5][0-9]'
        return re.compile(pattern).match(str(timedelta)).group()

    def _promote_process(self):
        pass


class ExcelHandleView(View):
    template_name = None

    FONT_NAME = '游ゴシック'

    TITLE_CELL = 'A2'
    USER_NAME_CELL = 'A4'

    SHEET_HEADLINE = '勤務報告書　{}'
    SHEET_USER_NAME = '氏名　{}'
    SHEET_TITLE = '一覧'
    SHEET_HEADER = ['日付', '曜日', '区分', '出勤', '退勤', '休憩開始', '休憩終了', '備考']

    DAY_COL_NUM = 1
    DOW_COL_NUM = 2
    DAY_KIND_COL_NUM = 3
    START_WORK_COL_NUM = 4
    END_WORK_COL_NUM = 5
    ENTER_BREAK_COL_NUM = 6
    END_BREAK_COL_NUM = 7
    ERR_MSG_COL_NUM = 9

    HEADER_ROW_NUM = 5

    WHITE = 'FFFFFF'
    BLACK = '000000'
    YELLOW = 'FFFF00'
    RED = 'FF0000'
    LIGHT_BLUE = '9BB3D4'
    LIGHT_PINK = 'E0B9B8'
    LIGHT_GREEN = 'DBE3C0'

    def _adjust_col_width_ws(self, ws):
        for column in ws.columns:
            column_letter = column[0].column_letter
            # ヘッダー以下の文字数に応じて幅を調整する
            max_char_length = max([len(str(cell.value)) for cell in column[self.HEADER_ROW_NUM - 1:] if cell.value])

            ws.column_dimensions[column_letter].width = max_char_length * 1.5 + 2

    def _edit_appearance_ws(self, ws):
        ws.merge_cells(start_row=self.HEADER_ROW_NUM, start_column=self.DOW_COL_NUM,
                       end_row=self.HEADER_ROW_NUM, end_column=self.DAY_KIND_COL_NUM)
        side = Side(style='thin', color=self.BLACK)

        for row in ws:
            day_kind = row[self.DAY_KIND_COL_NUM - 1].value
            DOW = row[self.DOW_COL_NUM - 1].value
            if day_kind in ['休日', '祝日']:
                fill_color = PatternFill(patternType='solid', fgColor=self._get_cell_color(day_kind, DOW))
                row_num = row[0].row
                ws[ws.cell(row=row_num, column=self.DAY_COL_NUM).coordinate].fill = fill_color
                ws[ws.cell(row=row_num, column=self.DOW_COL_NUM).coordinate].fill = fill_color

            for cell in row:
                ws[cell.coordinate].font = Font(name=self.FONT_NAME, size=9)

                if cell.value == '祝日':
                    ws[cell.coordinate].font = Font(color=self.RED, name=self.FONT_NAME, size=9)

                if cell.coordinate in (self.TITLE_CELL, self.USER_NAME_CELL):
                    ws[cell.coordinate].font = Font(name=self.FONT_NAME, size=12)
                    ws.row_dimensions[cell.row].height = 20

                if cell.row == self.HEADER_ROW_NUM:
                    fill_color = PatternFill(patternType='solid', fgColor=self.LIGHT_GREEN)
                    ws[cell.coordinate].fill = fill_color

                if cell.row < self.HEADER_ROW_NUM:
                    continue

                # ヘッダー以下を中央揃え、罫線を引く
                ws[cell.coordinate].border = Border(top=side, bottom=side, left=side, right=side)
                ws[cell.coordinate].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

    def _get_cell_color(self, day_kind, DOW):
        if day_kind == '祝日' or DOW == '日':
            return self.LIGHT_PINK
        elif DOW == '土':
            return self.LIGHT_BLUE


def get_DOW(datetime) -> str:
    week = {1: '月', 2: '火', 3: '水', 4: '木', 5: '金', 6: '土', 7: '日'}
    DOW_int = datetime.isoweekday()
    return week[DOW_int]


def get_work_days_by_qs(stamps_qs) -> list:
    work_days = [stamp_qs['datetime'].day for stamp_qs in stamps_qs.annotate(
        datetime=TruncDay('stamped_time')).values('datetime').distinct()]
    return work_days


def get_daily_stamps_info(stamps_qs, date):
    daily_stamps_qs = stamps_qs.filter(
        stamped_time__gte=date, stamped_time__lt=date + relativedelta(days=1))

    start_work_obj = daily_stamps_qs.filter(kind=TimeCard.Kind.IN).first()
    end_work_obj = daily_stamps_qs.filter(kind=TimeCard.Kind.OUT).first()
    enter_break_obj = daily_stamps_qs.filter(kind=TimeCard.Kind.ENTER_BREAK).first()
    end_break_obj = daily_stamps_qs.filter(kind=TimeCard.Kind.END_BREAK).first()

    start_work = start_work_obj.stamped_time if start_work_obj else ''
    end_work = end_work_obj.stamped_time if end_work_obj else ''
    enter_break = enter_break_obj.stamped_time if enter_break_obj else ''
    end_break = end_break_obj.stamped_time if end_break_obj else ''

    return start_work, end_work, enter_break, end_break


def calculation_hours_daily(start_work, end_work, enter_break, end_break) -> (timedelta, timedelta):
    try:
        work_hours = break_hours = timedelta()
        if start_work and end_work and start_work < end_work:
            work_hours = end_work - start_work

        if enter_break and end_break and enter_break < end_break:
            break_hours = end_break - enter_break

        if break_hours > timedelta() and work_hours > break_hours:
            work_hours -= break_hours

        return work_hours, break_hours
    except:
        return timedelta(), timedelta()


def timedelta2str(timedelta) -> str:
    seconds = timedelta.days * (24 * 60 * 60) + timedelta.seconds
    hours_float = float('{:.2f}'.format(seconds / (60 * 60)))
    return str(hours_float)


def get_toast_msg_by_session(session) -> dict:
    context = {}
    toast_key_list = ['success', 'error', 'warning']
    for key in toast_key_list:
        if session.get(key):
            context[key] = session.pop(key)

    return context
